from contextlib import contextmanager
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from xagent.core.tools.core import excel as excel_module
from xagent.core.tools.core.excel import (
    calculate_cell_size_from_font,
    read_excel_cells,
    update_excel_cells,
)


@pytest.fixture
def temp_excel_file(tmp_path):
    """Create a temporary Excel file for testing"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    # Ensure using default worksheet and rename to "Sheet1"
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Test Data 1"
    ws["B2"] = "Test Data 2"
    ws["C3"] = 123
    wb.save(file_path)
    return str(file_path)


@pytest.mark.asyncio
async def test_read_excel_cells(temp_excel_file):
    """Test reading Excel cell content"""
    result = await read_excel_cells(temp_excel_file)
    assert len(result) == 3
    assert any("A1: Test Data 1" in cell for cell in result)
    assert any("B2: Test Data 2" in cell for cell in result)
    assert any("C3: 123" in cell for cell in result)


@pytest.mark.asyncio
async def test_read_excel_cells_empty_sheet(tmp_path):
    """Test reading empty Excel file"""
    file_path = tmp_path / "empty.xlsx"
    wb = Workbook()
    # Ensure using default worksheet and rename to "Sheet1"
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(file_path)
    result = await read_excel_cells(str(file_path))
    assert len(result) == 1
    assert "No non-empty cells were found" in result[0]


@pytest.mark.asyncio
async def test_read_excel_cells_file_not_found():
    """Test reading non-existent file"""
    with pytest.raises(ValueError) as exc_info:
        await read_excel_cells("nonexistent.xlsx")
    assert "File not found" in str(exc_info.value)


@pytest.mark.asyncio
async def test_update_excel_cells_basic(temp_excel_file):
    """Test batch updating Excel cells (basic)"""
    updates = [
        {"cell_address": "D4", "new_value": "New Data 1"},
        {"cell_address": "E5", "new_value": "New Data 2"},
    ]
    result = await update_excel_cells(
        file_path=temp_excel_file,
        updates=updates,
        font_size=12,
        auto_size=True,
    )
    assert "Successfully batch updated 2 cells" in result

    # Verify updated content
    updated_content = await read_excel_cells(temp_excel_file)
    assert any("D4: New Data 1" in cell for cell in updated_content)
    assert any("E5: New Data 2" in cell for cell in updated_content)


@pytest.mark.asyncio
async def test_update_excel_cells_guards_the_entire_file_mutation(
    temp_excel_file, monkeypatch
):
    guarded_paths = []

    @contextmanager
    def guard_path(path):
        guarded_paths.append(Path(path).resolve())
        yield Path(path).resolve()

    monkeypatch.setattr(
        excel_module,
        "GLOBAL_PATH_MUTATION_LOCKS",
        SimpleNamespace(guard_path=guard_path),
        raising=False,
    )

    result = await update_excel_cells(
        file_path=temp_excel_file,
        updates=[{"cell_address": "A1", "new_value": "Guarded"}],
    )

    assert "Successfully batch updated 1 cells" in result
    assert guarded_paths == [Path(temp_excel_file).resolve()]


def test_calculate_cell_size_from_font():
    """Test cell size calculation function"""
    # Test basic calculation
    result = calculate_cell_size_from_font(font_size=12, text="Test Text", padding=2.0)
    assert isinstance(result, dict)
    assert "row_height" in result
    assert "column_width" in result

    # Test auto-wrap vs no-wrap
    long_text = "This is a very long test text that needs auto-wrap display"
    result_nowrap = calculate_cell_size_from_font(
        font_size=12, text=long_text, max_width=50, wrap_text=False
    )
    result_wrap = calculate_cell_size_from_font(
        font_size=12, text=long_text, max_width=50, wrap_text=True
    )
    assert result_wrap["row_height"] > result_nowrap["row_height"]


@pytest.mark.asyncio
async def test_update_excel_cells_invalid_cell(temp_excel_file):
    """Test updating invalid cell address (empty)"""
    updates = [{"cell_address": "", "new_value": "Test Data"}]
    result = await update_excel_cells(file_path=temp_excel_file, updates=updates)
    assert "Successfully batch updated 0 cells" in result


# -------------------- New tests --------------------


@pytest.mark.asyncio
async def test_invalid_sheet_name_read(temp_excel_file):
    with pytest.raises(ValueError) as exc_info:
        await read_excel_cells(temp_excel_file, sheet_name="NotExist")
    assert "Error reading Excel file" in str(exc_info.value)


@pytest.mark.asyncio
async def test_invalid_sheet_name_update(temp_excel_file):
    updates = [{"cell_address": "A1", "new_value": "X"}]
    with pytest.raises(ValueError) as exc_info:
        await update_excel_cells(
            file_path=temp_excel_file,
            updates=updates,
            sheet_name="NotExist",
        )
    assert "Batch update Excel failed" in str(exc_info.value)


@pytest.mark.asyncio
async def test_update_excel_cells_hyperlinks(temp_excel_file):
    updates = [
        {
            "cell_address": "A1",
            "new_value": "Visit Google",
            "hyperlink": "https://www.google.com",
        },
        {"cell_address": "B2", "new_value": "Go to A1", "hyperlink": "#Sheet1!A1"},
    ]
    result = await update_excel_cells(
        file_path=temp_excel_file, updates=updates, font_size=12
    )
    assert "Successfully batch updated 2 cells" in result

    wb = load_workbook(filename=temp_excel_file)
    ws = wb["Sheet1"]
    # External link
    cell_a1 = ws["A1"]
    assert cell_a1.value == "Visit Google"
    assert cell_a1.hyperlink is not None
    assert getattr(cell_a1.hyperlink, "target", None) == "https://www.google.com"
    assert cell_a1.font.underline == "single"
    # Internal link
    cell_b2 = ws["B2"]
    assert cell_b2.value == "Go to A1"
    assert cell_b2.hyperlink is not None
    assert getattr(cell_b2.hyperlink, "target", None) == "#Sheet1!A1"


@pytest.mark.asyncio
async def test_update_excel_cells_comments(temp_excel_file):
    updates = [
        {
            "cell_address": "C3",
            "new_value": "Has Comment",
            "comment_text": "This is a comment",
        },
        {
            "cell_address": "D4",
            "new_value": "Has Author",
            "comment_text": "Check this",
            "comment_author": "John",
        },
    ]
    result = await update_excel_cells(
        file_path=temp_excel_file, updates=updates, font_size=12
    )
    assert "Successfully batch updated 2 cells" in result

    wb = load_workbook(filename=temp_excel_file)
    ws = wb["Sheet1"]
    c3 = ws["C3"]
    assert c3.value == "Has Comment"
    assert c3.comment is not None
    assert c3.comment.text == "This is a comment"
    # default author
    assert c3.comment.author == "Gemini"

    d4 = ws["D4"]
    assert d4.value == "Has Author"
    assert d4.comment is not None
    assert d4.comment.text == "Check this"
    assert d4.comment.author == "John"
