from io import BytesIO
from pathlib import Path
from unittest.mock import Mock, patch

import pytest
from openpyxl import Workbook

from xagent.providers.pdf_parser.base import ParseResult
from xagent.providers.pdf_parser.deepdoc import DeepDocParser, _parse_xlsx_rows


# A fixture to easily access test resource files
@pytest.fixture
def resource_path() -> Path:
    """Provides the correct path to the test resources directory."""
    return Path("tests/resources/test_files")


# List of files for the smoke test
# We exclude files that deepdoc is not expected to handle, like .ppt, .jpg
SUPPORTED_FILES = [
    "test.pdf",
    "test.docx",
    "test.xlsx",
    "test.csv",
    "test.html",
    "test.json",
    "test.md",
    "test.txt",
    "gbk.txt",
]


@pytest.mark.asyncio
async def test_deepdoc_raw_output_deepdoc_engine(resource_path: Path):
    """Test DeepDoc parser with raw output enabled using DeepDoc engine."""
    test_file = resource_path / "test.pdf"
    parser = DeepDocParser(enable_raw_output=True)

    result = await parser.parse(str(test_file))

    # Check basic structure
    assert result is not None
    assert isinstance(result, ParseResult)

    # Check raw output fields
    assert result.raw_parser_output is not None
    assert result.parser_engine == "deepdoc"
    assert result.has_visualization_data

    # Check raw output structure
    raw_output = result.raw_parser_output
    assert "format" in raw_output
    assert raw_output["format"] == "deepdoc_pdf"
    assert "bboxes" in raw_output
    assert "total_elements" in raw_output
    assert "has_positions" in raw_output

    # Check visualization elements
    viz_elements = result.get_visualization_elements()
    assert isinstance(viz_elements, list)

    if viz_elements:  # Only check structure if elements exist
        element = viz_elements[0]
        required_keys = ["id", "type", "content", "bbox", "page", "metadata"]
        for key in required_keys:
            assert key in element

        # Check bbox structure
        bbox = element["bbox"]
        assert all(k in bbox for k in ["x0", "y0", "x1", "y1"])

        # Check metadata structure
        metadata = element["metadata"]
        assert "parser" in metadata
        assert "content_type" in metadata
        assert "has_image" in metadata
        assert metadata["parser"] == "deepdoc"

    # Check visualization summary
    viz_summary = result.get_visualization_summary()
    assert viz_summary["available"] is True
    assert viz_summary["parser_engine"] == "deepdoc"
    assert "total_elements" in viz_summary
    assert "elements_by_type" in viz_summary


@pytest.mark.asyncio
async def test_deepdoc_standard_mode(resource_path: Path):
    """Test DeepDoc parser in standard mode (no raw output)."""
    test_file = resource_path / "test.pdf"
    parser = DeepDocParser(enable_raw_output=False)

    result = await parser.parse(str(test_file))

    # Check basic structure
    assert result is not None
    assert isinstance(result, ParseResult)

    # Check that raw output is disabled
    assert result.raw_parser_output is None
    assert result.parser_engine is None
    assert not result.has_visualization_data

    # Check that visualization elements returns empty list
    viz_elements = result.get_visualization_elements()
    assert viz_elements == []


@pytest.mark.asyncio
@pytest.mark.parametrize("filename", SUPPORTED_FILES)
async def test_deepdoc_smoke_test(resource_path: Path, filename: str):
    """Ensures the DeepDocParser can process all supported file types without crashing."""
    test_file = resource_path / filename
    parser = DeepDocParser()

    try:
        result = await parser.parse(str(test_file))
        assert result is not None
        assert isinstance(result, ParseResult)
        # A basic check that some content was produced
        assert result.text_segments or result.figures or result.tables
    except Exception as e:
        # If gbk.txt fails with a UnicodeDecodeError, we accept that as a known limitation for now
        if "gbk" in filename and isinstance(e, UnicodeDecodeError):
            pytest.skip(f"DeepDocParser (TxtParser) does not support GBK encoding: {e}")
        else:
            pytest.fail(
                f"DeepDocParser failed on {filename} with an unexpected exception: {e}"
            )


@pytest.mark.asyncio
async def test_deepdoc_pdf_details(resource_path: Path):
    """Tests detailed, lossless extraction from a PDF file."""
    test_file = resource_path / "test.pdf"
    parser = DeepDocParser()
    # Pass a mock doc_id for the image saving path
    result = await parser.parse(str(test_file), doc_id="test_doc_id")

    assert result.text_segments, "No text segments were extracted from the PDF."
    assert result.tables, "No tables were extracted from the PDF."

    if result.figures:
        figure = result.figures[0]
        assert isinstance(figure.text, str)
        assert figure.metadata.get("type") == "figure"
        figure_image_path_str = figure.metadata.get("image_path")
        assert isinstance(figure_image_path_str, str) and figure_image_path_str
        assert Path(figure_image_path_str).exists(), (
            f"Figure image file not found at {figure_image_path_str}"
        )

    # Deep inspect the first table
    table = result.tables[0]
    assert isinstance(table.html, str) and table.html
    assert table.metadata.get("type") == "table"
    table_image_path_str = table.metadata.get("image_path")
    assert isinstance(table_image_path_str, str) and table_image_path_str
    assert Path(table_image_path_str).exists(), (
        f"Table image file not found at {table_image_path_str}"
    )

    # Deep inspect the first figure if present

    # Check that basic metadata is present (position info is now in raw_parser_output)
    first_segment = result.text_segments[0]
    assert "layout_type" in first_segment.metadata, "Layout type should be in metadata"
    assert "doc_id" in first_segment.metadata, "Doc ID should be in metadata"
    assert "page_number" in first_segment.metadata, "Page number should be in metadata"

    # Check layout metadata for text segments - accept both 'text' and 'title'
    layout_type = first_segment.metadata.get("layout_type")
    assert layout_type in ["text", "title"], (
        f"Text segments should have layout_type 'text' or 'title', got '{layout_type}'"
    )

    # Check basic metadata for tables
    if result.tables:
        table = result.tables[0]
        assert table.metadata.get("layout_type") == "table", (
            "Tables should have layout_type 'table'"
        )

        # Check basic metadata for figures
    if result.figures:
        figure = result.figures[0]
        assert figure.metadata.get("layout_type") == "figure", (
            "Figures should have layout_type 'figure'"
        )


@pytest.mark.asyncio
async def test_deepdoc_docx_details(resource_path: Path):
    """Tests detailed, lossless extraction from a DOCX file."""
    test_file = resource_path / "test.docx"
    parser = DeepDocParser()
    result = await parser.parse(str(test_file))

    assert result.text_segments, "No text segments were extracted from the DOCX."

    # Check for style metadata
    first_segment = result.text_segments[0]
    assert "style" in first_segment.metadata, "Style metadata is missing."

    # Check for tables if any
    if result.tables:
        assert result.tables[0].html is not None


@pytest.mark.asyncio
async def test_deepdoc_excel_details(resource_path: Path):
    """Tests detailed extraction from an XLSX file."""
    test_file = resource_path / "test.xlsx"
    parser = DeepDocParser()
    result = await parser.parse(str(test_file))

    assert result.text_segments, "No text segments were extracted from the XLSX."
    # Check that rows are parsed as segments
    assert len(result.text_segments) > 1
    # Check content of a known row/cell if possible (highly dependent on test file content)
    # For now, we just ensure segments are created.


@pytest.mark.asyncio
async def test_deepdoc_xlsx_parses_rows_without_repeating_title(tmp_path: Path):
    file_path = tmp_path / "structured.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Quarterly Enrollment Review"])
    ws.append(["Track", "Candidate ID", "Name", "Status"])
    ws.append(["Direct Entry", "A-0001", "Student One", "Passed"])
    ws.append(["Direct Entry", "A-0002", "Student Two", "Passed"])
    wb.save(file_path)

    parser = DeepDocParser()
    result = await parser.parse(str(file_path))

    assert result.text_segments[0].metadata["row_type"] == "title"
    assert result.text_segments[0].text == "Quarterly Enrollment Review"
    assert result.text_segments[1].metadata["row_type"] == "header"
    assert result.text_segments[1].text == "Track | Candidate ID | Name | Status"

    first_data_row = result.text_segments[2]
    assert first_data_row.metadata["row_type"] == "data"
    assert "Quarterly Enrollment Review" not in first_data_row.text
    assert (
        first_data_row.text
        == "Track: Direct Entry | Candidate ID: A-0001 | Name: Student One | Status: Passed"
    )


def test_parse_xlsx_rows_closes_workbook() -> None:
    workbook = Mock()
    workbook.sheetnames = ["Sheet1"]
    worksheet = Mock()
    worksheet.title = "Sheet1"
    worksheet.iter_rows.return_value = iter(
        [
            ("Quarterly Enrollment Review", None),
            ("Track", "Status"),
            ("Direct Entry", "Passed"),
        ]
    )
    workbook.worksheets = [worksheet]

    with patch(
        "xagent.providers.pdf_parser.deepdoc.load_workbook", return_value=workbook
    ):
        result = _parse_xlsx_rows("structured.xlsx")

    assert result.text_segments
    workbook.close.assert_called_once()


@pytest.mark.asyncio
async def test_deepdoc_invalid_xlsx_raises_parse_error() -> None:
    parser = DeepDocParser()

    with pytest.raises(ValueError, match="Failed to parse spreadsheet rows"):
        await parser.parse(BytesIO(b"not-a-valid-xlsx"), file_ext=".xlsx")
